import xlsxwriter 
from calendar import monthrange
import os.path
import datetime
import pandas as pd
#import the writer
import xlwt
#import the reader
import xlrd
import openpyxl


class DataBase:

    def __init__(self):
        
    
    
        datee = datetime.datetime.now()

        years = datee.year
        months = datee.strftime("%B")
        days = datee.day
        num_days = monthrange(years, datee.month)[1] # num_days = 28

        filename = str(years)+".xlsx"
        self.filename = filename
        self.sheet = months
        self.num_days = num_days
        self.days = days
        self.sheet2 = "Workers_details"
    # Workbook() takes one, non-optional, argument  
    # which is the filename that we want to create. 
        if os.path.isfile(self.filename):
            workbook = xlsxwriter.Workbook(self.filename)
            worksheet = workbook.get_worksheet_by_name(self.sheet)
            if worksheet is None:
                worksheet = workbook.add_worksheet(self.sheet)
                worksheet.write(0, 0, 'Name')
                row = 0
                column = 1
                for i in range(self.num_days):
                    #print(i+1)
                    worksheet.write(row,column,i+1)
                    column += 1

            
            df = pd.read_excel(self.filename, sheet_name=self.sheet)

            print("Column headings:")
            print(df.columns)

            
        else:
            workbook = xlsxwriter.Workbook(self.filename) 
            worksheet = workbook.add_worksheet(self.sheet) 
            worksheet.write(0, 0, "Name") 
            row = 0
            column = 1
            for i in range(self.num_days):
                #print(i+1)
                worksheet.write(row,column,i+1)
                column += 1

            worksheet2 = workbook.add_worksheet(self.sheet2)
            worksheet2.write(0,0,"Name")
            worksheet2.write(0,1,"Address")
            worksheet2.write(0,2,"Contact")
            worksheet2.write(0,3,"Salary")
            worksheet2.write(0,4,"Condition")
            # Finally, close the Excel file 
            # via the close() method. 
            workbook.close()

        
        
    def add_users(self, name, address, contact, salary):
        wb = openpyxl.load_workbook(filename = self.filename)     
        ws = wb.get_sheet_by_name(self.sheet2)
        #df = pd.read_excel(self.filename, sheet_name=self.sheet2)
        #row = len(df) + 2
        row = ws.max_row + 1
        wcell_name = ws.cell(row,1)
        wcell_name.value = name
        wcell_address = ws.cell(row,2)
        wcell_address.value = address
        wcell_contact = ws.cell(row,3)
        wcell_contact.value = contact
        wcell_salary = ws.cell(row,4)
        wcell_salary.value = salary
        wcell_condition = ws.cell(row,5)
        wcell_condition.value = "Active"

        ws2 = wb.get_sheet_by_name(self.sheet)
        rows = ws2.max_row + 1
        sheet1_name = ws2.cell(rows,1)
        sheet1_name.value = name
        wb.save(self.filename)

    def count_rows(self):
        wb = openpyxl.load_workbook(filename = self.filename)     
        ws = wb.get_sheet_by_name(self.sheet2)
        #df = pd.read_excel(self.filename, sheet_name=self.sheet2)
        #row = len(df) + 2
        row = ws.max_row
        return row

    def view_users(self,row):
        wb = openpyxl.load_workbook(filename = self.filename)     
        ws = wb.get_sheet_by_name(self.sheet2)
        self.users = {}
        #row = ws.max_row
        wcell_name = ws.cell(row,1)
        #print(wcell_name.value)
        #wcell_name.value = name
        wcell_address = ws.cell(row,2)
        #wcell_address.value = address
        wcell_contact = ws.cell(row,3)
        #wcell_contact.value = contact
        wcell_salary = ws.cell(row,4)
        #wcell_salary.value = salary
        wcell_condition = ws.cell(row,5)
        self.users[row] = (wcell_name.value,wcell_address.value, wcell_contact.value, wcell_salary.value)
        return self.users[row]

        
